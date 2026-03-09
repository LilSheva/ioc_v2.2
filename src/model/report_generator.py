"""Модуль для генерации отчетов IOC с поддержкой режимов ФСТЕК и ГосСОПКА."""

import re
import os
from typing import Dict, List, Any, Tuple
from urllib.parse import urlparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ReportGenerator:
    """Генератор отчетов IOC."""

    def __init__(self, ioc_config: List[Dict[str, Any]], uri_clean_mode: str = "unique"):
        """
        Инициализация генератора отчетов.

        Args:
            ioc_config: Конфигурация типов IOC
            uri_clean_mode: Режим очистки URI - "unique" (до уникального префикса) или "domain" (до домена)
        """
        self.ioc_config = ioc_config
        self.uri_clean_mode = uri_clean_mode
    
    def _smart_clean_uri(self, uris: List[Tuple[str, str, dict]]) -> Dict[str, str]:
        """
        Очистка URI: сокращает до уникального префикса или до домена (зависит от режима).

        Returns: cleaned_uri -> display_uri
        """
        domain_groups = {}
        for original, cleaned, metadata in uris:
            try:
                parsed = urlparse(cleaned if cleaned.startswith('http') else 'http://' + cleaned)
                domain = parsed.netloc or parsed.path.split('/')[0]
                if domain not in domain_groups:
                    domain_groups[domain] = []
                domain_groups[domain].append(cleaned)
            except:
                domain_groups[cleaned] = [cleaned]

        if self.uri_clean_mode == "domain":
            cleaned_map = {}
            for domain, uri_list in domain_groups.items():
                for uri in uri_list:
                    cleaned_map[uri] = domain
            return cleaned_map

        # Режим "unique" - сокращаем до уникального префикса
        cleaned_map = {}
        for domain, uri_list in domain_groups.items():
            # Если домен — IP-адрес, всегда сокращаем до домена
            if self._is_ip_address(domain):
                for uri in uri_list:
                    cleaned_map[uri] = domain
                continue

            if len(uri_list) == 1:
                cleaned_map[uri_list[0]] = domain
            else:
                # Сравниваем по сегментам пути, а не через startswith
                path_segments = {}
                for uri in uri_list:
                    try:
                        parsed = urlparse(uri if uri.startswith('http') else 'http://' + uri)
                        segments = [s for s in parsed.path.strip('/').split('/') if s]
                        path_segments[uri] = segments
                    except:
                        path_segments[uri] = []

                for uri in uri_list:
                    segments = path_segments[uri]
                    if not segments:
                        cleaned_map[uri] = domain
                        continue
                    # Находим минимальную глубину, отличающую от остальных
                    cleaned = domain
                    for depth in range(1, len(segments) + 1):
                        prefix = '/'.join(segments[:depth])
                        is_unique = all(
                            other_uri == uri or
                            '/'.join(path_segments[other_uri][:depth]) != prefix
                            for other_uri in uri_list
                        )
                        if is_unique:
                            cleaned = domain + '/' + prefix
                            break
                    cleaned_map[uri] = cleaned

        return cleaned_map
    
    def generate_xlsx_report(self, ioc_data: Dict[str, List[Tuple[str, str, dict]]],
                            output_path: str, bulletin: str = "", mode: str = "fstek",
                            event_type: str = "Фишинговая рассылка электронной почты. Вредоносные вложения") -> bool:
        """
        Генерирует форматированный .xlsx отчет с 10 столбцами.

        Args:
            ioc_data: Данные IOC в формате Dict[str, List[Tuple[original, cleaned, metadata]]]
            output_path: Путь для сохранения отчета
            bulletin: Бюллетень (для режима ФСТЕК)
            mode: Режим работы - "fstek" или "gossopka"
            event_type: Тип события по умолчанию (для режима ФСТЕК)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "IOC Report"
            
            headers = [
                "№", "Дата\nОтчёта", "Статус\nАктивности\nNTA",
                "Статус\nАктивности\nSIEM (Tools)", "Статус\nАктивности\nSIEM (MP)",
                "Тип\nИндикатора", "Индикатор", "IOC", "Бюллетень", "Тип события"
            ]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            uri_smart_map = {}
            if 'URI' in ioc_data:
                uri_smart_map = self._smart_clean_uri(ioc_data['URI'])

            row_num = 2
            counter = 1
            today_date = datetime.now().strftime('%d.%m.%Y')

            for ioc_config in self.ioc_config:
                if not ioc_config.get('enabled', False):
                    continue

                name = ioc_config['name']
                if name in ioc_data and ioc_data[name]:
                    report_type = ioc_config['report_type']
                    nta_status = ioc_config['nta_status']
                    siem_tools_status = ioc_config['siem_tools_status']
                    siem_status = ioc_config['siem_status']

                    for original, cleaned, metadata in ioc_data[name]:
                        display_original = original
                        if name == 'File':
                            display_original = re.sub(r'\s+', ' ', original)

                        ioc_display = cleaned
                        if name == 'URI' and cleaned in uri_smart_map:
                            ioc_display = uri_smart_map[cleaned]

                        if mode == "gossopka":
                            bulletin_num = metadata.get("bulletin_num", "")
                            file_bulletin = f"GosSOPKA {bulletin_num}" if bulletin_num else metadata.get("filename", "")
                            file_event_type = metadata.get("event_type", event_type)
                        else:
                            file_bulletin = bulletin
                            file_event_type = event_type

                        row_data = [
                            counter, today_date, nta_status, siem_tools_status,
                            siem_status, report_type, display_original,
                            ioc_display, file_bulletin, file_event_type
                        ]
                        ws.append(row_data)
                        counter += 1
                        row_num += 1

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            default_alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

            for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        if cell.column_letter in ('A', 'B', 'C', 'D', 'E', 'F'):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = default_alignment
                        
                        if cell.column_letter == 'A':
                            cell.fill = header_fill
                            cell.font = header_font

            ws.column_dimensions['A'].width = 4
            ws.column_dimensions['B'].width = 11
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 90
            ws.column_dimensions['H'].width = 90
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 80

            wb.save(output_path)
            return True

        except Exception as e:
            print(f"Ошибка при генерации .xlsx отчета: {e}")
            return False

    def generate_query_data(self, ioc_data: Dict[str, List[Tuple[str, str, dict]]]) -> List[Dict[str, Any]]:
        """
        Генерирует структурированные данные запросов для отображения в GUI.

        URI предобрабатываются: домены → DNS, IP-адреса → IP. Отдельная группа URI не создаётся.
        """
        # --- Предобработка URI: извлечь домены/IP, объединить с DNS/IP ---
        merged_data: Dict[str, List[Tuple[str, str, dict]]] = {}
        for key, items in ioc_data.items():
            if items:
                merged_data[key] = list(items)

        if 'URI' in merged_data and merged_data['URI']:
            uri_items = merged_data.pop('URI')
            for original, cleaned, metadata in uri_items:
                try:
                    parsed = urlparse(cleaned if cleaned.startswith('http') else 'http://' + cleaned)
                    domain = parsed.netloc or parsed.path.split('/')[0]
                except Exception:
                    domain = cleaned

                if self._is_ip_address(domain):
                    merged_data.setdefault('IP', []).append((original, domain, metadata))
                else:
                    merged_data.setdefault('DNS', []).append((original, domain, metadata))

        # Дедупликация cleaned-значений внутри каждой группы
        for key in merged_data:
            seen = set()
            deduped = []
            for item in merged_data[key]:
                if item[1] not in seen:
                    seen.add(item[1])
                    deduped.append(item)
            merged_data[key] = deduped

        # --- Собираем шаблоны по типам (DNS + URI шаблоны объединяются для DNS) ---
        config_by_name = {cfg['name']: cfg for cfg in self.ioc_config if cfg.get('enabled', False)}

        # Маппинг: query-группа -> список ioc_config names, из которых брать шаблоны
        type_template_sources: Dict[str, List[str]] = {}
        for cfg in self.ioc_config:
            if not cfg.get('enabled', False):
                continue
            name = cfg['name']
            if name == 'URI':
                # URI шаблоны объединяются с DNS
                type_template_sources.setdefault('DNS', []).append(name)
            else:
                type_template_sources.setdefault(name, []).append(name)

        # --- Генерация query_data ---
        query_data = []

        for ioc_config in self.ioc_config:
            if not ioc_config.get('enabled', False):
                continue

            name = ioc_config['name']
            if name == 'URI':
                continue  # URI не создаёт отдельную группу

            if name not in merged_data or not merged_data[name]:
                continue

            cleaned_iocs = [cleaned for _, cleaned, _ in merged_data[name]]

            # Собираем шаблоны из всех источников для этого типа
            sources = type_template_sources.get(name, [name])
            mp10_templates = []
            nad_templates = []
            for src_name in sources:
                src_cfg = config_by_name.get(src_name, {})
                mp10_templates.extend(src_cfg.get('mp10_templates', []))
                nad_templates.extend(src_cfg.get('nad_templates', []))
            mp10_templates = list(dict.fromkeys(mp10_templates))
            nad_templates = list(dict.fromkeys(nad_templates))

            group_queries = []

            for template in mp10_templates:
                group_queries.append({
                    'ioc_name': name, 'system': 'MP10',
                    'query': self._build_query(template, cleaned_iocs, " OR "),
                    'template': template,
                    'join_op': ' OR ',
                    'completed': False
                })

            for template in nad_templates:
                group_queries.append({
                    'ioc_name': name, 'system': 'NAD',
                    'query': self._build_query(template, cleaned_iocs, " || "),
                    'template': template,
                    'join_op': ' || ',
                    'completed': False
                })

            if group_queries:
                query_data.append({
                    'group_name': f"{name} ({ioc_config['report_type']})",
                    'ioc_count': len(cleaned_iocs),
                    'cleaned_iocs': cleaned_iocs,
                    'queries': group_queries
                })

        return query_data

    def generate_filters_xlsx(self, ioc_data: Dict[str, List[Tuple[str, str, dict]]],
                             output_path: str, log_callback=None) -> bool:
        """
        Генерирует файл "Фильтры.xlsx" программно (без шаблона).

        Создаёт листы: IP-адрес, DNS, Файл, E-mail, SHA256, SHA1, MD5.
        Дедупликация: IP из URI попадают на лист IP-адрес, все значения уникальны.

        Args:
            ioc_data: Данные IOC в формате Dict[str, List[Tuple[original, cleaned, metadata]]]
            output_path: Путь для сохранения нового файла фильтров
            log_callback: Функция для логирования (опционально)

        Returns:
            True если успешно, False при ошибке
        """
        def log(message):
            if log_callback:
                log_callback(message)
            else:
                print(message)

        try:
            sheet_mapping = {
                'IP': 'IP-адрес',
                'DNS': 'DNS',
                'URI': 'DNS',
                'Email': 'E-mail',
                'SHA256': 'SHA256',
                'SHA1': 'SHA1',
                'MD5': 'MD5',
                'File': 'Файл',
            }

            # Собираем данные по листам (set для дедупликации)
            sheet_data: Dict[str, set] = {}

            log(f"   • Обработка IOC данных...")
            for ioc_type, ioc_list in ioc_data.items():
                if not ioc_list:
                    log(f"      - {ioc_type}: пусто")
                    continue

                log(f"      - {ioc_type}: {len(ioc_list)} записей")
                for original, cleaned, metadata in ioc_list:
                    target_sheet = sheet_mapping.get(ioc_type)
                    filter_value = cleaned

                    if ioc_type == 'URI':
                        try:
                            parsed = urlparse(cleaned if cleaned.startswith('http') else 'http://' + cleaned)
                            domain = parsed.netloc or parsed.path.split('/')[0]
                            filter_value = domain

                            if self._is_ip_address(filter_value):
                                target_sheet = 'IP-адрес'
                                log(f"        URI {cleaned} -> IP-адрес ({filter_value})")
                        except:
                            filter_value = cleaned

                    elif ioc_type == 'DNS':
                        filter_value = cleaned.replace('[.]', '.').replace('[', '').replace(']', '')

                    elif ioc_type == 'Email':
                        filter_value = cleaned.replace('[.]', '.').replace('[', '').replace(']', '')

                    elif ioc_type == 'IP':
                        filter_value = cleaned.replace('[.]', '.').replace('[', '').replace(']', '')

                    if target_sheet:
                        if target_sheet not in sheet_data:
                            sheet_data[target_sheet] = set()
                        sheet_data[target_sheet].add(filter_value)

            log(f"   • Распределение по листам:")
            for sheet_name, ioc_set in sheet_data.items():
                log(f"      - {sheet_name}: {len(ioc_set)} уникальных записей")

            # Создаём книгу с нуля
            sheet_order = ['IP-адрес', 'DNS', 'Файл', 'E-mail', 'SHA256', 'SHA1', 'MD5']
            sheet_titles = {
                'IP-адрес': 'IP',
                'DNS': 'DNS',
                'Файл': 'File',
                'E-mail': 'E-mail',
                'SHA256': 'SHA256',
                'SHA1': 'SHA1',
                'MD5': 'MD5',
            }

            # Собираем шаблоны запросов по листам
            # Обратный маппинг: sheet_name -> list of ioc_config names
            sheet_to_types: Dict[str, List[str]] = {}
            for ioc_type, sname in sheet_mapping.items():
                sheet_to_types.setdefault(sname, []).append(ioc_type)

            sheet_templates: Dict[str, Dict[str, List[str]]] = {}  # sheet -> {mp10: [...], nad: [...]}
            for sname, ioc_types in sheet_to_types.items():
                mp10 = []
                nad = []
                for cfg in self.ioc_config:
                    if cfg['name'] in ioc_types:
                        mp10.extend(cfg.get('mp10_templates', []))
                        nad.extend(cfg.get('nad_templates', []))
                mp10 = list(dict.fromkeys(mp10))
                nad = list(dict.fromkeys(nad))
                sheet_templates[sname] = {'mp10': mp10, 'nad': nad}

            from openpyxl.utils import get_column_letter

            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFF")
            mp10_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            nad_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_font = Font(bold=True)
            center_al = Alignment(horizontal="center", vertical="center")

            log(f"   • Запись IOC в листы...")
            for sheet_name in sheet_order:
                ws = wb.create_sheet(title=sheet_name)

                tpl = sheet_templates.get(sheet_name, {})
                mp10_tpls = tpl.get('mp10', [])
                nad_tpls = tpl.get('nad', [])
                mp10_count = len(mp10_tpls)
                nad_count = len(nad_tpls)

                # A1 — заголовок типа
                ws['A1'] = sheet_titles[sheet_name]
                ws['A1'].font = title_font
                ws.column_dimensions['A'].width = 45

                # MP10 заголовок — объединённые ячейки (оранжевый)
                col_offset = 2  # столбец B
                if mp10_count > 0:
                    mp10_start = get_column_letter(col_offset)
                    mp10_end = get_column_letter(col_offset + mp10_count - 1)
                    if mp10_count > 1:
                        ws.merge_cells(f'{mp10_start}1:{mp10_end}1')
                    cell = ws.cell(row=1, column=col_offset, value='MP10')
                    cell.font = header_font
                    cell.fill = mp10_fill
                    cell.alignment = center_al
                    # Стиль для всех ячеек merged-диапазона
                    for c in range(col_offset, col_offset + mp10_count):
                        ws.cell(row=1, column=c).fill = mp10_fill
                        ws.column_dimensions[get_column_letter(c)].width = 45

                # NAD заголовок — объединённые ячейки (синий)
                nad_offset = col_offset + mp10_count
                if nad_count > 0:
                    nad_start = get_column_letter(nad_offset)
                    nad_end = get_column_letter(nad_offset + nad_count - 1)
                    if nad_count > 1:
                        ws.merge_cells(f'{nad_start}1:{nad_end}1')
                    cell = ws.cell(row=1, column=nad_offset, value='NAD')
                    cell.font = header_font
                    cell.fill = nad_fill
                    cell.alignment = center_al
                    for c in range(nad_offset, nad_offset + nad_count):
                        ws.cell(row=1, column=c).fill = nad_fill
                        ws.column_dimensions[get_column_letter(c)].width = 45

                # Данные: IOC в A, подставленные запросы в остальных столбцах
                ioc_values = sorted(sheet_data.get(sheet_name, set()))
                for idx, ioc in enumerate(ioc_values, start=2):
                    ws.cell(row=idx, column=1, value=ioc)
                    for ti, tmpl in enumerate(mp10_tpls):
                        ws.cell(row=idx, column=col_offset + ti,
                                value=tmpl.replace('{ioc}', ioc) + " OR")
                    for ti, tmpl in enumerate(nad_tpls):
                        ws.cell(row=idx, column=nad_offset + ti,
                                value=tmpl.replace('{ioc}', ioc) + " ||")

                log(f"      ✓ {sheet_name}: записано {len(ioc_values)} записей")

            log(f"   • Сохранение файла: {output_path}")
            wb.save(output_path)
            log(f"   ✓ Файл фильтров успешно создан")
            return True

        except Exception as e:
            print(f"Ошибка при генерации фильтров: {e}")
            return False

    @staticmethod
    def _build_query(template: str, ioc_values: List[str], join_op: str) -> str:
        """
        Формирует запрос из шаблона и списка IOC.

        Извлекает имя поля из шаблона и формирует field in ["v1", "v2", ...].
        """
        # Попытка извлечь имя поля: всё до оператора перед "{ioc}"
        m = re.match(r'^(.+?)\s*(?:={1,2}|!=|<>|CONTAINS|contains|LIKE|like|~|IN|in)\s*"\{ioc\}"$', template)
        if not m:
            # Мягкий fallback: всё до пробела(ов) перед "{ioc}"
            m = re.match(r'^(.+?)\s+"\{ioc\}"$', template)
        if m:
            field = m.group(1).rstrip()
            values_str = ", ".join(f'"{v}"' for v in ioc_values)
            return f'{field} in [{values_str}]'
        # Последний fallback — если шаблон совсем нестандартный
        queries = [template.replace('{ioc}', ioc) for ioc in ioc_values]
        return join_op.join(queries)

    def _is_ip_address(self, value: str) -> bool:
        """Проверяет, является ли строка IP-адресом."""
        try:
            import socket
            socket.inet_aton(value)
            return True
        except:
            return False