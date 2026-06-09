"""
Адаптер для генерации отчетов Excel (openpyxl).
"""

import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from ioc_analyzer.core.constants import DEFAULT_FSTEC_EVENT_TYPE
from ioc_analyzer.core.models import ReportData
from ioc_analyzer.core.parser.cleaner import smart_clean_uri


def generate_xlsx_report(
    report_data: ReportData,
    output_path: str,
    ioc_config: list[dict],
    uri_clean_mode: str = "domain",
) -> bool:
    """
    Генерирует форматированный .xlsx отчет с 10 столбцами.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "IOC Report"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "№", "Дата\nОтчёта", "Статус\nАктивности\nNTA",
            "Статус\nАктивности\nSIEM (Tools)", "Статус\nАктивности\nSIEM (MP)",
            "Тип\nИндикатора", "Индикатор", "IOC", "Бюллетень", "Тип события"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Вытаскиваем все URI
        uris = [ioc for ioc in report_data.indicators if ioc.ioc_type == 'URI']
        uri_smart_map = smart_clean_uri(uris, uri_clean_mode)

        row_num = 2
        counter = 1
        today_date = datetime.now().strftime('%d.%m.%Y')

        # Сначала сгруппируем индикаторы по типам в порядке конфигурации
        ioc_by_type = {}
        for ioc in report_data.indicators:
            ioc_by_type.setdefault(ioc.ioc_type, []).append(ioc)

        for cfg in ioc_config:
            if not cfg.get('enabled', False):
                continue

            name = cfg['name']
            if name in ioc_by_type and ioc_by_type[name]:
                report_type = cfg['report_type']
                nta_status = cfg['nta_status']
                siem_tools_status = cfg['siem_tools_status']
                siem_status = cfg['siem_status']

                for ioc in ioc_by_type[name]:
                    display_original = ioc.raw_value
                    if name == 'File':
                        display_original = re.sub(r'\s+', ' ', ioc.raw_value)

                    ioc_display = ioc.clean_value
                    if name == 'URI' and ioc.clean_value in uri_smart_map:
                        ioc_display = uri_smart_map[ioc.clean_value]

                    row_mode = ioc.parser_mode or report_data.parser_mode
                    if row_mode == "gossopka":
                        file_bulletin = f"GosSOPKA {ioc.source_file}" if ioc.source_file else report_data.source_filename
                        file_event_type = ioc.context or ""
                    else:
                        file_bulletin = report_data.source_filename
                        file_event_type = ioc.context or DEFAULT_FSTEC_EVENT_TYPE

                    row_data = [
                        counter, today_date, nta_status, siem_tools_status,
                        siem_status, report_type, display_original,
                        ioc_display, file_bulletin, file_event_type
                    ]
                    ws.append(row_data)
                    counter += 1
                    row_num += 1

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        default_alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

        for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    if cell.column_letter in ('A', 'B', 'C', 'D', 'E', 'F'):
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = default_alignment
                    
                    if cell.column_letter == 'A':
                        cell.fill = header_fill
                        cell.font = header_font

        ws.row_dimensions[1].height = 35
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 11
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 90
        ws.column_dimensions['H'].width = 90
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 80

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Ошибка при генерации xlsx отчета: {e}")
        return False


def generate_cve_xlsx_report(bdu_list: list[str], output_path: str) -> bool:
    """
    Генерирует Excel-таблицу для CVE/BDU с заданным форматированием.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "CVE"
        ws.views.sheetView[0].showGridLines = True

        headers = ["№", "ID ППТС", "BDU", "CVSS", "Продукт", "Ссылка"]
        ws.append(headers)

        header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
        num_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        header_font = Font(bold=True)
        num_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        row_num = 2
        list_to_write = bdu_list if bdu_list else [""]
        for i, bdu in enumerate(list_to_write, 1):
            row_data = [i, "", bdu, "", "", ""]
            ws.append(row_data)
            row_num += 1

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    if cell.column_letter == 'A':
                        cell.fill = num_fill
                        cell.font = num_font
                        cell.alignment = center_alignment
                    elif cell.column_letter in ('B', 'C', 'D'):
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment

        ws.row_dimensions[1].height = 25
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 50

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Ошибка при создании CVE xlsx: {e}")
        return False
