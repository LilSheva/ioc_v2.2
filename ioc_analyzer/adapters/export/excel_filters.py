"""
Адаптер для генерации Excel-файла фильтров поисковых запросов.
"""

import logging
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ioc_analyzer.core.models import ReportData
from ioc_analyzer.core.parser.cleaner import is_ip_address
from ioc_analyzer.core.query_builder import build_query

logger = logging.getLogger("ioc_analyzer.excel_filters")


def generate_filters_xlsx(
    report_data: ReportData,
    output_path: str,
    ioc_config: list[dict]
) -> bool:
    """
    Генерирует Excel-файл Фильтры.xlsx с поисковыми запросами для MP10 и NAD.
    """
    try:
        sheet_mapping = {
            'IP': 'IP-адрес',
            'DNS': 'DNS',
            'URI': 'DNS',
            'Email': 'E-mail',
            'SHA256': 'SHA256',
            'SHA1': 'SHA1',
            'MD5': 'MD5',
            'File': 'Файл',
        }

        # Распределяем IOC по листам
        sheet_data: dict[str, set[str]] = {}
        for ioc in report_data.indicators:
            target_sheet = sheet_mapping.get(ioc.ioc_type)
            filter_value = ioc.clean_value

            if ioc.ioc_type == 'URI':
                try:
                    parsed = urlparse(
                        ioc.clean_value if ioc.clean_value.startswith('http') 
                        else 'http://' + ioc.clean_value
                    )
                    domain = parsed.netloc or parsed.path.split('/')[0]
                    filter_value = domain
                    if is_ip_address(filter_value):
                        target_sheet = 'IP-адрес'
                except Exception:
                    filter_value = ioc.clean_value

            if target_sheet:
                sheet_data.setdefault(target_sheet, set()).add(filter_value)

        sheet_order = ['IP-адрес', 'DNS', 'Файл', 'E-mail', 'SHA256', 'SHA1', 'MD5']
        sheet_titles = {
            'IP-адрес': 'IP', 'DNS': 'DNS', 'Файл': 'File',
            'E-mail': 'E-mail', 'SHA256': 'SHA256', 'SHA1': 'SHA1', 'MD5': 'MD5',
        }

        # Собираем шаблоны для каждого листа
        sheet_templates = {}
        for sname in sheet_order:
            mp10 = []
            nad = []
            # Ищем типы IOC, привязанные к этому листу
            assoc_types = [k for k, v in sheet_mapping.items() if v == sname]
            for cfg in ioc_config:
                if cfg['name'] in assoc_types:
                    mp10.extend(cfg.get('mp10_templates', []))
                    nad.extend(cfg.get('nad_templates', []))
            sheet_templates[sname] = {
                'mp10': list(dict.fromkeys(mp10)),
                'nad': list(dict.fromkeys(nad))
            }

        wb = Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист

        header_font = Font(bold=True, color="FFFFFF")
        mp10_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        nad_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True)
        center_al = Alignment(horizontal="center", vertical="center")

        for sheet_name in sheet_order:
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            tpls = sheet_templates[sheet_name]
            mp10_tpls = tpls['mp10']
            nad_tpls = tpls['nad']
            mp10_count = len(mp10_tpls)
            nad_count = len(nad_tpls)

            ws['A1'] = sheet_titles[sheet_name]
            ws['A1'].font = title_font
            ws.column_dimensions['A'].width = 45

            col_offset = 2
            # Секция MP10
            if mp10_count > 0:
                mp10_start = get_column_letter(col_offset)
                mp10_end = get_column_letter(col_offset + mp10_count - 1)
                if mp10_count > 1:
                    ws.merge_cells(f'{mp10_start}1:{mp10_end}1')
                cell = ws.cell(row=1, column=col_offset, value='MP10')
                cell.font = header_font
                cell.fill = mp10_fill
                cell.alignment = center_al
                for c in range(col_offset, col_offset + mp10_count):
                    ws.cell(row=1, column=c).fill = mp10_fill
                    ws.column_dimensions[get_column_letter(c)].width = 45

            # Секция NAD
            nad_offset = col_offset + mp10_count
            if nad_count > 0:
                nad_start = get_column_letter(nad_offset)
                nad_end = get_column_letter(nad_offset + nad_count - 1)
                if nad_count > 1:
                    ws.merge_cells(f'{nad_start}1:{nad_end}1')
                cell = ws.cell(row=1, column=nad_offset, value='NAD')
                cell.font = header_font
                cell.fill = nad_fill
                cell.alignment = center_al
                for c in range(nad_offset, nad_offset + nad_count):
                    ws.cell(row=1, column=c).fill = nad_fill
                    ws.column_dimensions[get_column_letter(c)].width = 45

            # Записываем данные
            ioc_values = sorted(list(sheet_data.get(sheet_name, set())))
            for idx, ioc in enumerate(ioc_values, start=2):
                ws.cell(row=idx, column=1, value=ioc)
                for ti, tmpl in enumerate(mp10_tpls):
                    ws.cell(
                        row=idx, 
                        column=col_offset + ti,
                        value=tmpl.replace('{ioc}', ioc) + " OR"
                    )
                for ti, tmpl in enumerate(nad_tpls):
                    ws.cell(
                        row=idx, 
                        column=nad_offset + ti,
                        value=tmpl.replace('{ioc}', ioc) + " ||"
                    )

        wb.save(output_path)
        return True
    except Exception as e:
        logger.error("Ошибка при генерации фильтров: %s", e, exc_info=True)
        return False
